from __future__ import annotations

import hashlib
import json
import re
import uuid
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Callable
from io import BytesIO

import polars as pl
from openpyxl import load_workbook
from pypdf import PdfReader

from .config import settings
from .database import (
    COMPANY_ID,
    current_baseline_version,
    get_duckdb,
    get_sqlite,
    load_mapping_profile,
    next_data_version,
    save_mapping_profile,
    utc_now,
)
from .pipeline import CONTEXT_ONLY_DATASETS, DEPENDENCIES, refresh_context_layers, refresh_gold_layers
from .document_routing import strong_filename_document_hint

ProgressCallback = Callable[[str, int, str], None]


COMPLETE_UPLOAD_STATUSES = {"committed", "stored_source", "pending_mapping"}
TYPE_ALIASES = {
    "invoices": "supplier_invoices",
    "assets": "balance_sheet",
    "liabilities": "balance_sheet",
    "assets_liabilities": "balance_sheet",
    "transactions": "bank_statements",
    "budgets": "sales_forecast",
    "contracts": "material_contracts",
    "generic": "business_requirements",
}


def _canonical_document_type(value: str | None) -> str:
    key = str(value or "").strip().lower()
    return TYPE_ALIASES.get(key, key)


def _existing_upload_types(row: Any) -> set[str]:
    values = {_canonical_document_type(row["document_type"] if "document_type" in row.keys() else "")}
    try:
        metadata = json.loads(row["metadata_json"] or "{}")
    except Exception:
        metadata = {}
    for item in metadata.get("detected_document_types") or []:
        values.add(_canonical_document_type(str(item)))
    declared = str(row["declared_document_type"] or "") if "declared_document_type" in row.keys() else ""
    if declared and declared != "auto":
        values.add(_canonical_document_type(declared))
    return {item for item in values if item}


def _prepare_existing_duplicate(row: Any, intake_category: str, declared_document_type: str) -> tuple[Any | None, bool, bool]:
    """Return the surviving duplicate row, or None when a corrective retry should proceed.

    Failed/stale attempts must never poison the content hash. A user may also
    deliberately re-upload identical bytes with an explicit corrected document
    type; in that case the old contribution is removed and rebuilt from source.
    Category-only corrections are applied without duplicating business rows.
    """
    status = str(row["processing_status"] or "").strip().lower()
    requested_type = _canonical_document_type(declared_document_type)
    correction_requested = (
        declared_document_type != "auto"
        and requested_type
        and requested_type not in _existing_upload_types(row)
    )
    stale_attempt = status not in COMPLETE_UPLOAD_STATUSES and status != "processing"
    if stale_attempt or correction_requested:
        upload_id = int(row["id"])
        from .data_management import purge_uploaded_file_for_reprocessing
        purge_uploaded_file_for_reprocessing(upload_id)
        return None, True, False

    existing_category = str(row["intake_category"] or "recurring")
    category_changed = existing_category != intake_category
    if category_changed:
        from .data_management import move_uploaded_file
        move_uploaded_file(int(row["id"]), intake_category)
        sql = get_sqlite()
        try:
            row = sql.execute("SELECT * FROM uploaded_files WHERE id=?", (int(row["id"]),)).fetchone()
        finally:
            sql.close()
    return row, False, category_changed


def _emit_progress(callback: ProgressCallback | None, stage: str, progress: int, message: str) -> None:
    if callback:
        callback(stage, max(0, min(int(progress), 100)), message)

DOCUMENT_ALIASES: dict[str, dict[str, list[str]]] = {
    "transactions": {
        "transaction_date": ["transaction_date", "date", "posted_date", "booking_date"],
        "description": ["description", "narration", "details", "memo", "transaction_details"],
        "category": ["category", "type", "transaction_type"],
        "amount": ["amount", "value", "total", "net_amount"],
    },
    "bank_statements": {
        "transaction_date": ["transaction_date", "date", "posted_date", "value_date"],
        "description": ["description", "narration", "details", "memo"],
        "account_name": ["account_name", "account", "bank_account"],
        "amount": ["amount", "net_amount", "transaction_amount"],
        "debit": ["debit", "withdrawal", "money_out"],
        "credit": ["credit", "deposit", "money_in"],
        "balance": ["balance", "running_balance", "closing_balance"],
        "currency": ["currency", "currency_code"],
    },
    "payments": {
        "payment_date": ["payment_date", "date", "paid_date"],
        "reference": ["reference", "payment_reference", "invoice_number", "invoice_no"],
        "counterparty": ["counterparty", "supplier", "vendor", "customer", "payee"],
        "amount": ["amount", "paid_amount", "payment_amount", "total"],
        "currency": ["currency", "currency_code"],
        "status": ["status", "payment_status"],
    },
    "supplier_invoices": {
        "invoice_number": ["invoice_number", "invoice_no", "invoice", "bill_number", "bill_no"],
        "supplier": ["supplier", "vendor", "supplier_name", "vendor_name", "company"],
        "invoice_date": ["invoice_date", "bill_date", "date"],
        "due_date": ["due_date", "payment_due", "due"],
        "amount": ["amount", "total", "invoice_total", "gross_amount", "amount_due", "grand_total"],
        "subtotal": ["subtotal", "net_amount", "amount_ex_gst", "exclusive_amount"],
        "gst_amount": ["gst", "gst_amount", "tax", "tax_amount"],
        "description": ["description", "line_item", "details", "item_description", "memo"],
        "supplier_abn": ["supplier_abn", "abn", "vendor_abn"],
        "status": ["status", "payment_status"],
        "currency": ["currency", "currency_code"],
        "sku": ["sku", "product_code", "item_code", "stock_code"],
        "quantity": ["quantity", "qty", "units"],
        "unit_cost": ["unit_cost", "cost", "purchase_price", "unit_price"],
    },
    "sales_invoices": {
        "invoice_number": ["invoice_number", "invoice_no", "invoice"],
        "supplier": ["customer", "client", "customer_name", "client_name", "supplier", "company"],
        "invoice_date": ["invoice_date", "date"],
        "due_date": ["due_date", "payment_due", "due"],
        "amount": ["amount", "total", "invoice_total", "gross_amount", "amount_due", "grand_total"],
        "subtotal": ["subtotal", "net_amount", "amount_ex_gst", "exclusive_amount"],
        "gst_amount": ["gst", "gst_amount", "tax", "tax_amount"],
        "description": ["description", "line_item", "details", "item_description", "memo"],
        "supplier_abn": ["supplier_abn", "abn", "vendor_abn"],
        "status": ["status", "payment_status"],
        "currency": ["currency", "currency_code"],
        "sku": ["sku", "product_code", "item_code", "stock_code"],
        "quantity": ["quantity", "qty", "units"],
        "unit_cost": ["unit_cost", "cost", "purchase_price", "unit_price"],
    },
    "invoices": {
        "invoice_number": ["invoice_number", "invoice_no", "invoice", "bill_number"],
        "supplier": ["supplier", "vendor", "customer", "client", "company"],
        "invoice_date": ["invoice_date", "date"],
        "due_date": ["due_date", "payment_due"],
        "amount": ["amount", "total", "invoice_total", "amount_due", "grand_total"],
        "subtotal": ["subtotal", "net_amount", "amount_ex_gst"],
        "gst_amount": ["gst", "gst_amount", "tax", "tax_amount"],
        "description": ["description", "line_item", "details", "item_description", "memo"],
        "supplier_abn": ["supplier_abn", "abn", "vendor_abn"],
        "status": ["status", "payment_status"],
        "currency": ["currency", "currency_code"],
        "sku": ["sku", "product_code", "item_code", "stock_code"],
        "quantity": ["quantity", "qty", "units"],
        "unit_cost": ["unit_cost", "cost", "purchase_price", "unit_price"],
    },
    "assets": {
        "name": ["asset_name", "name", "account", "item", "description"],
        "category": ["asset_category", "category", "type"],
        "classification": ["classification", "term", "current_non_current"],
        "amount": ["current_value", "amount", "balance", "value", "purchase_cost"],
    },
    "liabilities": {
        "name": ["liability_name", "name", "account", "item", "description"],
        "category": ["liability_type", "category", "type"],
        "classification": ["classification", "term", "current_non_current", "maturity"],
        "amount": ["current_balance", "amount", "balance", "value", "original_amount"],
    },
    "assets_liabilities": {
        "name": ["name", "account", "account_name", "item", "description"],
        "category": ["category", "type", "account_type"],
        "classification": ["classification", "term", "current_non_current"],
        "amount": ["amount", "balance", "value"],
    },
    "customers": {
        "code": ["customer_code", "customer_id", "client_id", "code"],
        "name": ["customer", "customer_name", "client", "client_name", "name"],
        "country": ["country", "customer_country", "location"],
        "segment": ["segment", "customer_segment", "category"],
        "status": ["status", "active"],
    },
    "suppliers": {
        "code": ["supplier_code", "supplier_id", "vendor_id", "code"],
        "name": ["supplier", "supplier_name", "vendor", "vendor_name", "name"],
        "country": ["country", "supplier_country", "location"],
        "category": ["category", "supplier_category", "type"],
        "currency": ["currency", "currency_code"],
        "status": ["status", "active"],
    },
    "inventory": {
        "sku": ["sku", "product_code", "item_code", "stock_code"],
        "name": ["product", "product_name", "item", "description", "name"],
        "quantity": ["quantity", "qty", "stock_on_hand"],
        "unit_cost": ["unit_cost", "cost", "purchase_price"],
        "total_value": ["total_value", "inventory_value", "stock_value"],
        "location": ["location", "warehouse", "store"],
        "status": ["status", "stock_status"],
    },
    "budgets": {
        "period": ["period", "month", "date", "financial_period"],
        "account": ["account", "account_name", "line_item"],
        "category": ["category", "type"],
        "budget_amount": ["budget", "budget_amount", "planned", "forecast"],
        "actual_amount": ["actual", "actual_amount"],
    },
    "balance_sheet": {
        "period_start": ["period_start", "start_date"],
        "period_end": ["period_end", "end_date", "date"],
        "line_item": ["line_item", "account", "account_name", "item"],
        "amount": ["amount", "balance", "value"],
        "currency": ["currency", "currency_code"],
    },
    "profit_loss": {
        "period_start": ["period_start", "start_date"],
        "period_end": ["period_end", "end_date", "date"],
        "line_item": ["line_item", "account", "account_name", "item"],
        "amount": ["amount", "balance", "value"],
        "currency": ["currency", "currency_code"],
    },
    "cash_flow_statement": {
        "period_start": ["period_start", "start_date"],
        "period_end": ["period_end", "end_date", "date"],
        "line_item": ["line_item", "account", "account_name", "item"],
        "amount": ["amount", "balance", "value"],
        "currency": ["currency", "currency_code"],
    },
    "payroll": {
        "pay_period": ["pay_period", "period", "pay_run", "pay_date", "payment_date"],
        "employee": ["employee", "employee_name", "staff", "worker", "name"],
        "gross_pay": ["gross_pay", "gross", "gross_wages", "gross_earnings", "total_earnings"],
        "ordinary_time_earnings": ["ordinary_time_earnings", "ote", "ordinary_earnings", "superable_earnings"],
        "payg_withholding": ["payg_withholding", "payg", "tax_withheld", "withholding", "income_tax"],
        "superannuation": ["superannuation", "super", "sg", "employer_super", "super_contribution"],
        "net_pay": ["net_pay", "net", "take_home_pay", "net_wages"],
        "currency": ["currency", "currency_code"],
        "status": ["status", "payroll_status"],
    },
    "market_context": {
        "signal_type": ["signal_type", "type", "category"],
        "topic": ["topic", "event", "indicator", "title"],
        "entity": ["entity", "company", "competitor", "supplier", "commodity"],
        "geography": ["geography", "country", "region", "location"],
        "observed_at": ["observed_at", "observed_date", "date"],
        "published_at": ["published_at", "published_date"],
        "value": ["value", "signal_value", "price", "index_value"],
        "unit": ["unit", "currency"],
        "direction": ["direction", "trend", "movement"],
        "source_name": ["source_name", "source", "publisher"],
        "source_url": ["source_url", "url", "link"],
        "relevance_score": ["relevance_score", "relevance", "score"],
        "estimated_impact": ["estimated_impact", "impact", "business_impact"],
        "impact_horizon": ["impact_horizon", "horizon", "time_horizon"],
    },
}

REQUIRED_FIELDS: dict[str, list[str]] = {
    "transactions": ["description", "amount"], "bank_statements": ["description"], "payments": ["amount"],
    "supplier_invoices": ["invoice_number", "amount"], "sales_invoices": ["invoice_number", "amount"], "invoices": ["invoice_number", "amount"],
    "assets": ["name", "amount"], "liabilities": ["name", "amount"], "assets_liabilities": ["name", "amount"],
    "customers": ["name"], "suppliers": ["name"], "inventory": ["name"], "budgets": ["period", "budget_amount"],
    "balance_sheet": ["line_item", "amount"], "profit_loss": ["line_item", "amount"], "cash_flow_statement": ["line_item", "amount"],
    "payroll": ["gross_pay"], "market_context": ["topic"],
}


def normalise_name(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_") or "column"


def unique_names(names: list[str]) -> list[str]:
    seen: dict[str, int] = {}; result: list[str] = []
    for name in names:
        base = normalise_name(name); count = seen.get(base, 0); seen[base] = count + 1
        result.append(base if count == 0 else f"{base}_{count + 1}")
    return result



def _period_from_balance_header(value: Any) -> tuple[str, str] | None:
    """Convert a balance-sheet column heading into a reporting period."""
    if isinstance(value, datetime):
        end = value.date()
    elif isinstance(value, date):
        end = value
    else:
        text = str(value or "").strip()
        if not text:
            return None
        end = None
        for fmt in ("%d %b %Y", "%d %B %Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                end = datetime.strptime(text, fmt).date()
                break
            except ValueError:
                pass
        if end is None:
            match = re.search(r"(?:19|20)\d{2}", text)
            if not match:
                return None
            year = int(match.group(0))
            end = date(year, 12, 31)
    try:
        prior = end.replace(year=end.year - 1)
    except ValueError:
        prior = end.replace(year=end.year - 1, day=28)
    start = prior + timedelta(days=1)
    return start.isoformat(), end.isoformat()


def _extract_balance_sheet_template(rows: list[tuple[Any, ...]], sheet_title: str) -> pl.DataFrame | None:
    """Convert the business.gov.au-style wide balance sheet into LedgerFlow's long schema."""
    if not rows:
        return None
    preview = " ".join(str(value or "") for row in rows[:12] for value in row).lower()
    if "balance sheet" not in preview and "balance sheet" not in sheet_title.lower():
        return None

    label_col: int | None = None
    for col_index in range(max((len(row) for row in rows), default=0)):
        labels = {
            normalise_name(str(row[col_index]))
            for row in rows[:50]
            if col_index < len(row) and row[col_index] is not None
        }
        if "current_assets" in labels and ("total_assets" in labels or "fixed_assets" in labels):
            label_col = col_index
            break
    if label_col is None:
        return None

    header_row: int | None = None
    periods: list[tuple[int, str, str]] = []
    for row_index, row in enumerate(rows[:12]):
        candidates: list[tuple[int, str, str]] = []
        for col_index in range(label_col + 1, len(row)):
            period = _period_from_balance_header(row[col_index])
            if period:
                candidates.append((col_index, period[0], period[1]))
        if candidates:
            header_row = row_index
            periods = candidates
            break
    if header_row is None or not periods:
        return None

    section_aliases = {
        "current_assets": "current_asset",
        "fixed_assets": "non_current_asset",
        "non_current_assets": "non_current_asset",
        "current_short_term_liabilities": "current_liability",
        "current_liabilities": "current_liability",
        "long_term_liabilities": "non_current_liability",
        "non_current_liabilities": "non_current_liability",
    }
    excluded = {
        "total", "total_assets", "total_liabilities", "net_assets_net_worth",
        "net_assets", "working_capital", "assumptions",
    }
    records: list[dict[str, Any]] = []
    section: str | None = None
    for row in rows[header_row + 1:]:
        if label_col >= len(row):
            continue
        raw_label = row[label_col]
        label = str(raw_label or "").strip()
        key = normalise_name(label)
        if key in section_aliases:
            section = section_aliases[key]
            continue
        if not label or key in excluded or key.startswith("more") or section is None:
            continue
        for col_index, period_start, period_end in periods:
            value = row[col_index] if col_index < len(row) else None
            if value is None or str(value).strip() == "":
                continue
            amount = safe_float(value)
            if "liability" in section:
                amount = -abs(amount)
            records.append({
                "period_start": period_start,
                "period_end": period_end,
                "line_item": label,
                "amount": amount,
                "currency": "AUD",
                "classification": section,
            })
    return pl.DataFrame(records, strict=False) if records else None

def read_workbook(path: Path) -> dict[str, pl.DataFrame]:
    if path.suffix.lower() == ".csv":
        frame = pl.read_csv(path, infer_schema_length=3000, ignore_errors=True, try_parse_dates=True)
        frame.columns = unique_names(frame.columns)
        return {"CSV": frame}
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        result: dict[str, pl.DataFrame] = {}
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows: continue
            template_frame = _extract_balance_sheet_template(rows, sheet.title)
            if template_frame is not None:
                result[sheet.title] = template_frame
                continue
            title_key = normalise_name(sheet.title)
            if title_key.startswith("using_this_balance_sheet"):
                continue
            header_index = 0
            for idx, row in enumerate(rows[:10]):
                non_empty = sum(value is not None and str(value).strip() != "" for value in row)
                if non_empty >= 2:
                    header_index = idx; break
            headers = unique_names([str(value or f"column_{index + 1}") for index, value in enumerate(rows[header_index])])
            records = [dict(zip(headers, row)) for row in rows[header_index + 1:] if any(value is not None for value in row)]
            result[sheet.title] = pl.DataFrame(records, strict=False) if records else pl.DataFrame(schema=headers)
        return result
    raise ValueError("Only CSV, XLSX, and XLSM files are supported in this version.")


def read_spreadsheet(path: Path) -> pl.DataFrame:
    frames = read_workbook(path)
    return next(iter(frames.values()), pl.DataFrame())


def schema_signature(columns: list[str]) -> str:
    return hashlib.sha256("|".join(sorted(columns)).encode()).hexdigest()


def find_column(columns: list[str], aliases: list[str]) -> str | None:
    for alias in aliases:
        if alias in columns: return alias
    for column in columns:
        if any(alias in column for alias in aliases): return column
    return None


def suggested_mapping(columns: list[str], document_type: str) -> dict[str, str]:
    return {field: column for field, aliases in DOCUMENT_ALIASES.get(document_type, {}).items() if (column := find_column(columns, aliases))}


def detect_document_type(columns: list[str], filename: str = "", sheet_name: str = "") -> tuple[str, float, dict[str, str]]:
    names = set(columns); joined = " ".join(columns + [normalise_name(filename), normalise_name(sheet_name)])
    filename_hint = strong_filename_document_hint(filename)
    if filename_hint:
        return filename_hint, 0.99, suggested_mapping(columns, filename_hint)
    # Strong structural rules prevent generic balance columns from routing bank exports as balance sheets.
    if "bank_statement" in joined or {"debit", "credit", "balance"}.issubset(names) or {"withdrawal", "deposit", "balance"}.issubset(names):
        mapping = suggested_mapping(columns, "bank_statements")
        return "bank_statements", 0.96, mapping
    if "market_context" in joined or {"topic", "relevance_score", "estimated_impact"}.issubset(names):
        mapping = suggested_mapping(columns, "market_context")
        return "market_context", 0.96, mapping
    payroll_mapping = suggested_mapping(columns, "payroll")
    if "gross_pay" in payroll_mapping and ("employee" in payroll_mapping or "pay_period" in payroll_mapping) and any(
        field in payroll_mapping for field in ("payg_withholding", "superannuation", "net_pay")
    ):
        return "payroll", 0.96, payroll_mapping
    # Invoice tables commonly contain generic date/description/amount fields, so identify
    # their structural invoice markers before scoring generic transaction exports.
    invoice_mapping = suggested_mapping(columns, "invoices")
    has_invoice_number = "invoice_number" in invoice_mapping
    has_invoice_amount = "amount" in invoice_mapping
    has_invoice_party = "supplier" in invoice_mapping
    has_invoice_dates = "invoice_date" in invoice_mapping or "due_date" in invoice_mapping
    invoice_context = any(token in joined for token in ("invoice", "supplier", "vendor", "customer", "client", "bill"))
    if has_invoice_number and has_invoice_amount and (has_invoice_party or has_invoice_dates or invoice_context):
        sales_context = any(token in joined for token in ("sales_invoice", "customer_invoice", "accounts_receivable", "customer", "client"))
        supplier_context = any(token in joined for token in ("supplier_invoice", "vendor_invoice", "accounts_payable", "supplier", "vendor", "bill"))
        if sales_context and not supplier_context:
            doc_type = "sales_invoices"
        elif supplier_context and not sales_context:
            doc_type = "supplier_invoices"
        else:
            doc_type = "invoices"
        mapping = suggested_mapping(columns, doc_type)
        richness = sum(1 for field in ("supplier", "invoice_date", "due_date", "subtotal", "gst_amount", "description", "currency") if field in mapping)
        return doc_type, round(min(0.99, 0.93 + richness * 0.008), 2), mapping
    hints = {
        "bank_statements": ["bank_statement", "running_balance", "withdrawal", "deposit"],
        "payments": ["payment_reference", "paid_amount", "payment_date"],
        "supplier_invoices": ["supplier_invoice", "vendor", "bill_number", "bill_no"],
        "sales_invoices": ["sales_invoice", "customer_name", "client_name"],
        "customers": ["customer_master", "customer_code", "customer_id"],
        "suppliers": ["supplier_master", "supplier_code", "vendor_id"],
        "inventory": ["inventory", "stock_on_hand", "sku", "product_code"],
        "budgets": ["budget_amount", "planned", "budget"],
        "balance_sheet": ["balance_sheet", "assets", "equity"],
        "profit_loss": ["profit_loss", "income_statement", "revenue", "expenses"],
        "cash_flow_statement": ["cash_flow_statement", "operating_activities", "financing_activities"],
        "payroll": ["payroll", "pay_run", "gross_pay", "payg", "superannuation", "net_pay"],
        "chart_of_accounts": ["chart_of_accounts", "coa", "account_code", "account_type"],
        "business_requirements": ["business_requirements", "brd", "user_story", "acceptance_criteria"],
        "fixed_asset_register": ["fixed_asset_register", "asset_register", "depreciation", "useful_life"],
        "aged_debtors_creditors": ["aged_debtors", "aged_creditors", "ageing", "days_overdue"],
        "material_contracts": ["material_contract", "contract_value", "renewal_date", "counterparty"],
        "sales_forecast": ["sales_forecast", "forecast_revenue", "expected_orders"],
        "personnel_plan": ["personnel_plan", "headcount", "role", "salary"],
        "use_cases_user_stories": ["use_cases", "user_stories", "persona", "acceptance_criteria"],
        "historical_tax_returns": ["tax_return", "taxable_income", "assessment_year"],
        "market_context": ["market_context", "relevance_score", "business_impact", "source_url", "geopolitical"],
        "assets": ["asset_register", "asset_name", "purchase_cost", "depreciation"],
        "liabilities": ["liability_register", "liability_name", "lender", "maturity_date"],
        "transactions": ["transaction", "narration", "description", "amount"],
    }
    scores: dict[str, float] = {}
    for doc_type, tokens in hints.items():
        token_score = sum(1 for token in tokens if token in joined)
        mapping = suggested_mapping(columns, doc_type)
        required = REQUIRED_FIELDS.get(doc_type, [])
        required_score = sum(1 for field in required if field in mapping)
        scores[doc_type] = token_score * 1.5 + required_score * 2 + len(mapping) * 0.15
    # Generic invoice fallback and combined asset/liability fallback.
    if "invoice" in joined or {"invoice_number", "due_date"}.intersection(names): scores["invoices"] = 4 + len(suggested_mapping(columns, "invoices")) * 0.2
    if "asset" in joined and "liabil" in joined: scores["assets_liabilities"] = 6 + len(suggested_mapping(columns, "assets_liabilities")) * 0.2
    best = max(scores, key=scores.get) if scores else "generic"
    score = scores.get(best, 0)
    mapping = suggested_mapping(columns, best)
    required = REQUIRED_FIELDS.get(best, [])
    has_required = all(field in mapping for field in required)
    confidence = min(0.99, 0.35 + score / 12) if has_required else min(0.69, score / 12)
    if score < 2.2: return "generic", 0.2, {}
    return best, round(confidence, 2), mapping


def safe_float(value: Any) -> float:
    if value is None or value == "": return 0.0
    if isinstance(value, (int, float)): return float(value)
    text = str(value).strip()
    negative = text.startswith("(") and text.endswith(")")
    cleaned = re.sub(r"[^0-9.\-]", "", text)
    try:
        number = float(cleaned or 0)
        return -abs(number) if negative else number
    except ValueError: return 0.0


def safe_date(value: Any) -> str:
    if isinstance(value, datetime): return value.date().isoformat()
    if isinstance(value, date): return value.isoformat()
    text = str(value or "").strip()
    if not text: return datetime.now(timezone.utc).date().isoformat()
    return text[:10]


def canonical_json(row: dict[str, Any]) -> str:
    return json.dumps({str(k): (v.isoformat() if hasattr(v, "isoformat") else v) for k, v in sorted(row.items())}, ensure_ascii=False, sort_keys=True, default=str)


def _mapped(row: dict[str, Any], mapping: dict[str, str], key: str, default: Any = None) -> Any:
    source = mapping.get(key); return row.get(source, default) if source else default


def _business_key(document_type: str, row: dict[str, Any], mapping: dict[str, str], index: int) -> str:
    pick = lambda key: str(_mapped(row, mapping, key, "") or "").strip().lower()
    if document_type in {"supplier_invoices", "sales_invoices", "invoices"}: return f"{pick('invoice_number')}|{pick('supplier')}" or f"row-{index}"
    if document_type in {"transactions", "bank_statements"}: return f"{pick('transaction_date')}|{pick('description')}|{safe_float(_mapped(row,mapping,'amount',0)):.2f}"
    if document_type == "payments": return f"{pick('payment_date')}|{pick('reference')}|{pick('counterparty')}|{safe_float(_mapped(row,mapping,'amount',0)):.2f}"
    if document_type in {"assets", "liabilities", "assets_liabilities"}: return pick("name") or f"row-{index}"
    if document_type in {"customers", "suppliers"}: return pick("code") or pick("name") or f"row-{index}"
    if document_type == "inventory": return pick("sku") or pick("name") or f"row-{index}"
    if document_type == "budgets": return f"{pick('period')}|{pick('account')}|{pick('category')}"
    if document_type == "payroll": return f"{pick('pay_period')}|{pick('employee')}"
    if document_type in {"balance_sheet", "profit_loss", "cash_flow_statement"}: return f"{pick('period_end')}|{pick('line_item')}"
    if document_type == "market_context": return f"{pick('published_at')}|{pick('topic')}|{pick('entity')}|{pick('geography')}"
    return hashlib.sha1(canonical_json(row).encode()).hexdigest()


def _fingerprint_decision(document_type: str, business_key: str, row_hash: str) -> tuple[str, int, str]:
    sql = get_sqlite()
    existing = sql.execute(
        "SELECT row_hash, record_version, record_id FROM row_fingerprints WHERE company_id=? AND document_type=? AND business_key=? AND is_current=1 ORDER BY id DESC LIMIT 1",
        (COMPANY_ID, document_type, business_key),
    ).fetchone(); sql.close()
    stable_id = f"lf-{normalise_name(document_type)}-{hashlib.sha1(business_key.encode()).hexdigest()[:16]}"
    if not existing: return "new", 1, stable_id
    if existing["row_hash"] == row_hash: return "unchanged", int(existing["record_version"]), str(existing["record_id"])
    return "changed", int(existing["record_version"]) + 1, str(existing["record_id"] or stable_id)


def _save_fingerprint(document_type: str, business_key: str, row_hash: str, version: int, record_id: str, upload_id: int, sheet: str, row_number: int) -> None:
    sql = get_sqlite()
    sql.execute("UPDATE row_fingerprints SET is_current=0 WHERE company_id=? AND document_type=? AND business_key=? AND is_current=1", (COMPANY_ID, document_type, business_key))
    sql.execute(
        "INSERT INTO row_fingerprints(company_id, document_type, business_key, row_hash, record_version, is_current, source_file_id, source_sheet, source_row_number, record_id, created_at) VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?, ?, ?)",
        (COMPANY_ID, document_type, business_key, row_hash, version, upload_id, sheet, row_number, record_id, utc_now()),
    ); sql.commit(); sql.close()


def _insert_record(con: Any, document_type: str, record_id: str, row: dict[str, Any], mapping: dict[str, str], source_file: str, row_number: int, upload_id: int = 0) -> None:
    m = lambda key, default=None: _mapped(row, mapping, key, default)
    today = datetime.now(timezone.utc).date().isoformat()
    if document_type in {"assets", "liabilities", "assets_liabilities"}:
        forced = "asset" if document_type == "assets" else "liability" if document_type == "liabilities" else ""
        raw_category = str(m("category", forced or "asset") or forced or "asset").lower()
        category = forced or ("liability" if "liabil" in raw_category or "debt" in raw_category else "asset")
        raw_class = str(m("classification", "current") or "current").lower()
        classification = "non-current" if any(token in raw_class for token in ["non", "long", ">12"]) else "current"
        con.execute("INSERT OR REPLACE INTO assets_liabilities VALUES (?, ?, ?, ?, ?, ?, ?)", (record_id, str(m("name", f"Imported item {row_number}")), category, classification, safe_float(m("amount", 0)), "imported", source_file))
    elif document_type in {"supplier_invoices", "sales_invoices", "invoices"}:
        invoice_kind = "sales" if document_type == "sales_invoices" else "supplier"
        amount = safe_float(m("amount", 0))
        subtotal = safe_float(m("subtotal", 0))
        gst_amount = safe_float(m("gst_amount", 0))
        description = str(m("description", "") or "")
        con.execute(
            """INSERT OR REPLACE INTO invoices(
                id, invoice_number, supplier, invoice_date, due_date, amount, status, source_file,
                invoice_kind, currency, subtotal, gst_amount, description, supplier_abn,
                account_code, category, tax_code, categorisation_confidence, validation_status, source_upload_id,
                sku, quantity, unit_cost
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '', '', 'REVIEW', 0, 'needs_review', ?, ?, ?, ?)""",
            (
                record_id, str(m("invoice_number", f"IMPORTED-{row_number}")),
                str(m("supplier", "Unknown counterparty")), safe_date(m("invoice_date", today)),
                safe_date(m("due_date", today)), amount, str(m("status", "due") or "due").lower(),
                source_file, invoice_kind, str(m("currency", "AUD") or "AUD"), subtotal,
                gst_amount, description, str(m("supplier_abn", "") or ""), upload_id,
                str(m("sku", "") or ""), safe_float(m("quantity", 0)), safe_float(m("unit_cost", 0)),
            ),
        )
    elif document_type == "transactions":
        con.execute("INSERT OR REPLACE INTO transactions VALUES (?, ?, ?, ?, ?, ?, ?)", (record_id, safe_date(m("transaction_date", today)), str(m("description", f"Imported transaction {row_number}")), str(m("category", "imported")), safe_float(m("amount", 0)), "imported", source_file))
    elif document_type == "bank_statements":
        amount = safe_float(m("amount", 0)) or safe_float(m("credit", 0)) - safe_float(m("debit", 0))
        description = str(m("description", f"Bank entry {row_number}"))
        con.execute("INSERT OR REPLACE INTO bank_transactions VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)", (record_id, safe_date(m("transaction_date", today)), description, str(m("account_name", "Bank account")), amount, safe_float(m("balance", 0)), str(m("currency", "")), "imported", source_file))
        con.execute("INSERT OR REPLACE INTO transactions VALUES (?, ?, ?, ?, ?, ?, ?)", (record_id, safe_date(m("transaction_date", today)), description, "bank transaction", amount, "imported", source_file))
    elif document_type == "payments":
        amount = safe_float(m("amount", 0)); ref = str(m("reference", "")); counterparty = str(m("counterparty", "Unknown counterparty"))
        con.execute("INSERT OR REPLACE INTO payments VALUES (?, ?, ?, ?, ?, ?, ?, ?)", (record_id, safe_date(m("payment_date", today)), ref, counterparty, amount, str(m("currency", "")), str(m("status", "imported")), source_file))
        con.execute("INSERT OR REPLACE INTO transactions VALUES (?, ?, ?, ?, ?, ?, ?)", (record_id, safe_date(m("payment_date", today)), f"{counterparty} {ref}".strip(), "payment", -abs(amount), "imported", source_file))
    elif document_type == "customers":
        con.execute("INSERT OR REPLACE INTO customers VALUES (?, ?, ?, ?, ?, ?, ?)", (record_id, str(m("code", "")), str(m("name", f"Customer {row_number}")), str(m("country", "")), str(m("segment", "")), str(m("status", "active")), source_file))
    elif document_type == "suppliers":
        con.execute("INSERT OR REPLACE INTO suppliers VALUES (?, ?, ?, ?, ?, ?, ?, ?)", (record_id, str(m("code", "")), str(m("name", f"Supplier {row_number}")), str(m("country", "")), str(m("category", "")), str(m("currency", "")), str(m("status", "active")), source_file))
    elif document_type == "inventory":
        qty = safe_float(m("quantity", 0)); unit_cost = safe_float(m("unit_cost", 0)); total = safe_float(m("total_value", 0)) or qty * unit_cost
        con.execute("INSERT OR REPLACE INTO inventory VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)", (record_id, str(m("sku", "")), str(m("name", f"Item {row_number}")), qty, unit_cost, total, str(m("location", "")), str(m("status", "active")), source_file))
    elif document_type == "budgets":
        budget = safe_float(m("budget_amount", 0)); actual = safe_float(m("actual_amount", 0))
        con.execute("INSERT OR REPLACE INTO budgets VALUES (?, ?, ?, ?, ?, ?, ?, ?)", (record_id, str(m("period", "")), str(m("account", "")), str(m("category", "")), budget, actual, actual - budget, source_file))
    elif document_type == "payroll":
        gross = safe_float(m("gross_pay", 0)); ote = safe_float(m("ordinary_time_earnings", gross)) or gross
        payg = safe_float(m("payg_withholding", 0)); super_amount = safe_float(m("superannuation", 0)); net = safe_float(m("net_pay", 0))
        con.execute(
            "INSERT OR REPLACE INTO payroll_records VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (record_id, str(m("pay_period", today)), str(m("employee", f"Employee {row_number}")), gross, ote, payg, super_amount, net, str(m("currency", "AUD") or "AUD"), str(m("status", "imported") or "imported"), source_file),
        )
    elif document_type in {"balance_sheet", "profit_loss", "cash_flow_statement"}:
        con.execute("INSERT OR REPLACE INTO statement_snapshots VALUES (?, ?, ?, ?, ?, ?, ?, ?)", (record_id, document_type, safe_date(m("period_start", today)), safe_date(m("period_end", today)), str(m("line_item", f"Line {row_number}")), safe_float(m("amount", 0)), str(m("currency", "")), source_file))
    elif document_type == "market_context":
        relevance = safe_float(m("relevance_score", 0.5)); relevance = relevance / 100 if relevance > 1 else relevance
        con.execute("INSERT OR REPLACE INTO market_signals VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", (record_id, str(m("signal_type", "market")), str(m("topic", f"Market signal {row_number}")), str(m("entity", "")), str(m("geography", "")), safe_date(m("observed_at", today)), safe_date(m("published_at", today)), str(m("value", "")), str(m("unit", "")), str(m("direction", "")), str(m("source_name", "Uploaded market file")), str(m("source_url", "")), relevance, str(m("estimated_impact", "")), str(m("impact_horizon", "")), source_file))
    else:
        con.execute("INSERT OR REPLACE INTO generic_documents VALUES (?, ?, ?, ?, ?, ?)", (record_id, document_type, str(row.get(next(iter(row), ""), f"Record {row_number}")), canonical_json(row), source_file, row_number))


def import_known_rows(frame: pl.DataFrame, document_type: str, source_file: str, mapping: dict[str, str] | None = None, upload_id: int = 0, sheet_name: str = "CSV") -> tuple[dict[str, int], list[str]]:
    mapping = mapping or suggested_mapping(frame.columns, document_type)
    required = REQUIRED_FIELDS.get(document_type, [])
    missing = [field for field in required if not mapping.get(field)]
    if missing: return {"new": 0, "changed": 0, "unchanged": 0, "rejected": frame.height}, [f"Missing required mappings: {', '.join(missing)}"]
    counts = {"new": 0, "changed": 0, "unchanged": 0, "rejected": 0}; issues: list[str] = []
    con = get_duckdb()
    for index, row in enumerate(frame.iter_rows(named=True), start=1):
        try:
            key = _business_key(document_type, row, mapping, index)
            row_hash = hashlib.sha256(canonical_json(row).encode()).hexdigest()
            decision, version, record_id = _fingerprint_decision(document_type, key, row_hash)
            if decision == "unchanged": counts["unchanged"] += 1; continue
            _insert_record(con, document_type, record_id, row, mapping, source_file, index, upload_id)
            _save_fingerprint(document_type, key, row_hash, version, record_id, upload_id, sheet_name, index)
            counts[decision] += 1
        except Exception as exc:
            counts["rejected"] += 1; issues.append(f"Row {index}: {type(exc).__name__}: {exc}")
    con.close(); return counts, issues[:50]


def _write_bronze(file_id: str, filename: str, content: bytes, digest: str) -> tuple[Path, Path]:
    safe_filename = re.sub(r"[^A-Za-z0-9._-]", "_", Path(filename).name)
    bronze_dir = settings.data_path / "bronze" / COMPANY_ID / file_id
    bronze_dir.mkdir(parents=True, exist_ok=True)
    original = bronze_dir / safe_filename; original.write_bytes(content)
    metadata = bronze_dir / "metadata.json"
    metadata.write_text(json.dumps({"file_id": file_id, "company_id": COMPANY_ID, "original_filename": filename, "sha256": digest, "size_bytes": len(content), "uploaded_at": utc_now()}, indent=2), encoding="utf-8")
    # Compatibility raw copy.
    raw = settings.data_path / "raw" / f"{digest[:12]}_{safe_filename}"; raw.write_bytes(content)
    return original, metadata


def _write_silver(frame: pl.DataFrame, document_type: str, file_id: str, sheet_name: str) -> Path:
    destination = settings.data_path / "silver" / COMPANY_ID / document_type
    destination.mkdir(parents=True, exist_ok=True)
    safe_sheet = re.sub(r"[^A-Za-z0-9._-]", "_", sheet_name)[:80]
    path = destination / f"{file_id}_{safe_sheet}.parquet"
    frame.write_parquet(path, compression="zstd")
    return path


def _pdf_text(content: bytes) -> str:
    try:
        reader = PdfReader(BytesIO(content))
        return "\n".join((page.extract_text() or "") for page in reader.pages).strip()
    except Exception:
        return ""


def _ocr_pdf_text(content: bytes) -> tuple[str, str]:
    if not settings.ocr_enabled:
        return "", "disabled"
    try:
        import fitz  # PyMuPDF
        import pytesseract
        from PIL import Image

        if settings.tesseract_cmd.strip():
            pytesseract.pytesseract.tesseract_cmd = settings.tesseract_cmd.strip()
        document = fitz.open(stream=content, filetype="pdf")
        pages: list[str] = []
        scale = max(1.0, float(settings.ocr_dpi) / 72.0)
        matrix = fitz.Matrix(scale, scale)
        for index, page in enumerate(document):
            if index >= max(1, settings.ocr_max_pages):
                break
            pixmap = page.get_pixmap(matrix=matrix, alpha=False)
            image = Image.frombytes("RGB", (pixmap.width, pixmap.height), pixmap.samples)
            pages.append(pytesseract.image_to_string(image))
        document.close()
        text = "\n".join(pages).strip()
        return text, "tesseract" if text else "no_text_detected"
    except Exception as exc:
        return "", f"unavailable:{type(exc).__name__}"


def _pdf_money(text: str, labels: list[str]) -> float:
    for label in labels:
        pattern = rf"(?im)^\s*{label}\s*[:\-]?\s*(?:AUD|AU\$|\$)?\s*\(?([0-9][0-9,]*\.\d{{2}})\)?\s*$"
        matches = re.findall(pattern, text)
        if matches:
            return safe_float(matches[-1])
    return 0.0


def _pdf_date(text: str, labels: list[str]) -> str:
    for label in labels:
        match = re.search(rf"(?i){label}\s*[:\-]?\s*(\d{{1,2}}[/-]\d{{1,2}}[/-]\d{{2,4}}|\d{{4}}-\d{{2}}-\d{{2}}|\d{{1,2}}\s+[A-Za-z]{{3,9}}\s+\d{{4}})", text)
        if match:
            raw = match.group(1)
            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y", "%Y-%m-%d", "%d %b %Y", "%d %B %Y"):
                try:
                    return datetime.strptime(raw, fmt).date().isoformat()
                except ValueError:
                    pass
            return safe_date(raw)
    return datetime.now(timezone.utc).date().isoformat()


def _extract_invoice_from_pdf(text: str, filename: str, declared_document_type: str) -> tuple[str, dict[str, Any], float] | None:
    compact = re.sub(r"\s+", " ", text).strip()
    signal = f"{filename} {compact[:4000]}".lower()
    if not any(token in signal for token in ["invoice", "tax invoice", "amount due", "bill to", "invoice no"]):
        return None
    kind = "sales_invoices" if declared_document_type == "sales_invoices" or any(token in filename.lower() for token in ["sales_invoice", "customer_invoice", "issued_invoice"]) else "supplier_invoices"
    invoice_match = re.search(r"(?i)(?:invoice\s*(?:number|no\.?|#)|inv\s*#)\s*[:\-]?\s*([A-Z0-9][A-Z0-9\-/]{2,})", text)
    invoice_number = invoice_match.group(1).strip() if invoice_match else f"PDF-{hashlib.sha1(content_key(filename, text).encode()).hexdigest()[:10].upper()}"
    abn_match = re.search(r"(?i)\bABN\s*[:\-]?\s*([0-9 ]{11,14})", text)
    abn = re.sub(r"\s+", "", abn_match.group(1)) if abn_match else ""
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines() if line.strip()]
    ignored = ("tax invoice", "invoice", "amount due", "bill to", "ship to", "abn", "date", "page ")
    supplier = next((line for line in lines[:15] if 2 < len(line) < 90 and any(ch.isalpha() for ch in line) and not line.lower().startswith(ignored)), "Unknown counterparty")
    if kind == "sales_invoices":
        # For an issued sales invoice, the accounting counterparty is the customer in
        # the Bill To section rather than the issuer shown in the letterhead.
        for index, line in enumerate(lines):
            lowered = line.lower().rstrip(":")
            if lowered in {"bill to", "customer", "client"} and index + 1 < len(lines):
                candidate = lines[index + 1]
                if 2 < len(candidate) < 120:
                    supplier = candidate
                    break
            match = re.match(r"(?i)(?:bill to|customer|client)\s*:\s*(.{3,120})$", line)
            if match:
                supplier = match.group(1).strip()
                break
    total = _pdf_money(text, [r"grand\s+total", r"total\s+due", r"amount\s+due", r"invoice\s+total", r"total"])
    gst = _pdf_money(text, [r"gst", r"tax"])
    subtotal = _pdf_money(text, [r"subtotal", r"total\s+ex(?:cluding)?\s+gst", r"net\s+amount"])
    if total <= 0:
        amounts = [safe_float(value) for value in re.findall(r"(?:AUD|AU\$|\$)\s*([0-9][0-9,]*\.\d{2})", text)]
        total = max(amounts) if amounts else 0.0
    description_candidates = [line for line in lines if len(line) > 12 and not any(token in line.lower() for token in ["invoice", "subtotal", "total", "gst", "abn", "bank", "payment", "due date"])]
    description = " · ".join(description_candidates[:3])[:500]
    invoice_date = _pdf_date(text, [r"invoice\s+date", r"date"])
    due_date = _pdf_date(text, [r"due\s+date", r"payment\s+due"])
    # A prose document that merely mentions invoices must never become a $0
    # invoice. If no payable/receivable amount can be extracted, preserve it as
    # source evidence or route it to review instead of posting a fake record.
    if total <= 0:
        return None
    confidence = 0.92 if invoice_match else 0.74
    record = {
        "invoice_number": invoice_number,
        "supplier": supplier,
        "invoice_date": invoice_date,
        "due_date": due_date,
        "amount": total,
        "subtotal": subtotal,
        "gst_amount": gst,
        "description": description,
        "supplier_abn": abn,
        "status": "due",
        "currency": "AUD",
    }
    return kind, record, confidence


def content_key(filename: str, text: str) -> str:
    return f"{filename}|{text[:1000]}"


INVOICE_PDF_TYPES = {"supplier_invoices", "sales_invoices", "invoices"}
CONTEXTUAL_PDF_TYPES = {
    "business_requirements", "material_contracts", "use_cases_user_stories",
    "historical_tax_returns", "market_context", "contracts", "generic",
}


def _pdf_document_hint(filename: str, text: str, declared_document_type: str) -> str:
    if declared_document_type != "auto":
        return declared_document_type
    filename_hint = strong_filename_document_hint(filename)
    if filename_hint:
        return filename_hint
    signal = normalise_name(f"{filename} {text[:3500]}")
    hints = [
        ("business_requirements", ("business_requirements", "business_requirement", "brd", "functional_requirements")),
        ("use_cases_user_stories", ("use_cases", "user_stories", "user_story", "acceptance_criteria")),
        ("material_contracts", ("material_supplier_contract", "supplier_agreement", "material_contract", "contract_term")),
        ("historical_tax_returns", ("historical_tax_return", "tax_return", "assessment_year")),
        ("market_context", ("market_context", "competitor_metric", "market_risk", "market_opportunity")),
    ]
    for document_type, tokens in hints:
        if any(token in signal for token in tokens):
            return document_type
    if any(token in signal for token in ("tax_invoice", "invoice_number", "invoice_no", "amount_due", "total_due")):
        return "sales_invoices" if any(token in signal for token in ("sales_invoice", "customer_invoice", "issued_invoice")) else "supplier_invoices"
    return "generic"


def store_source_document(
    filename: str, content: bytes, intake_category: str = "recurring", declared_document_type: str = "auto",
    progress_callback: ProgressCallback | None = None,
) -> dict[str, Any]:
    """Preserve PDFs and automatically extract digitally generated invoices when possible."""
    _emit_progress(progress_callback, "identifying", 10, "Fingerprinting and identifying the document")
    digest = hashlib.sha256(content).hexdigest()
    file_id = f"file_{digest[:16]}"
    sql = get_sqlite()
    duplicate = sql.execute("SELECT * FROM uploaded_files WHERE sha256=?", (digest,)).fetchone()
    if duplicate:
        sql.close()
        duplicate, corrective_retry, category_changed = _prepare_existing_duplicate(duplicate, intake_category, declared_document_type)
        sql = get_sqlite()
        if duplicate:
            category_changed = str(duplicate["intake_category"] or "recurring") == intake_category
            sql.close()
            return {
                "duplicate": True,
                "category_changed": category_changed,
                "corrective_retry": corrective_retry,
                "upload_id": duplicate["id"],
                "filename": filename,
                "document_type": duplicate["document_type"],
                "intake_category": duplicate["intake_category"] or intake_category,
                "rows_imported": 0,
                "issues": ["This exact source document was already stored. No duplicate copy was created."],
                "needs_mapping": str(duplicate["mapping_status"] or "") == "needs_mapping",
                "columns": json.loads(duplicate["columns_json"] or "[]"),
                "processing": {"new": 0, "changed": 0, "unchanged": int(duplicate["row_count"] or 1), "rejected": 0},
                "data_version": int(duplicate["data_version"] or 0),
                "baseline_version": int(duplicate["baseline_version"] or 0),
                "storage": "Existing Bronze source evidence",
            }
        sql.close()
        sql = get_sqlite()

    _emit_progress(progress_callback, "preserving", 22, "Preserving the original PDF in the Bronze evidence layer")
    original, metadata_path = _write_bronze(file_id, filename, content, digest)
    _emit_progress(progress_callback, "extracting", 38, "Reading digital text and identifying invoice fields")
    text = _pdf_text(content)
    extraction_method = "digital_text" if text else "none"
    if len(text.strip()) < 80 and settings.ocr_enabled:
        _emit_progress(progress_callback, "ocr", 46, "Running local OCR on image-only PDF pages")
        ocr_text, ocr_status = _ocr_pdf_text(content)
        if len(ocr_text.strip()) > len(text.strip()):
            text = ocr_text
            extraction_method = "local_tesseract_ocr"
        elif extraction_method == "none":
            extraction_method = ocr_status
    hinted_type = _pdf_document_hint(filename, text, declared_document_type)
    allow_invoice_extraction = hinted_type in INVOICE_PDF_TYPES or (
        declared_document_type == "auto" and hinted_type == "generic"
    )
    extraction = _extract_invoice_from_pdf(text, filename, hinted_type) if text and allow_invoice_extraction else None
    detected_type = extraction[0] if extraction else hinted_type
    contextual_source = detected_type in CONTEXTUAL_PDF_TYPES and detected_type != "generic"
    metadata = {
        "file_metadata_path": str(metadata_path),
        "intake_category": intake_category,
        "declared_document_type": declared_document_type,
        "source_only": extraction is None,
        "text_extracted": bool(text),
        "extraction_method": extraction_method,
        "extraction_status": "invoice_extracted" if extraction else "context_source_stored" if contextual_source else "human_review_required",
        "text_preview": text[:3000],
    }
    cursor = sql.execute(
        "INSERT INTO uploaded_files(created_at, filename, sha256, document_type, rows_imported, curated_path, columns_json, mapping_status, company_id, file_id, original_filename, file_type, file_size, raw_path, bronze_path, silver_paths_json, processing_status, metadata_json, last_processed_at, intake_category, declared_document_type) VALUES (?, ?, ?, ?, 0, '', '[]', ?, ?, ?, ?, ?, ?, ?, ?, '[]', ?, ?, ?, ?, ?)",
        (utc_now(), filename, digest, detected_type, "mapped" if extraction or contextual_source else "source_only", COMPANY_ID, file_id, filename, Path(filename).suffix.lower(), len(content), str(original), str(original.parent), "processing" if extraction else "stored_source", json.dumps(metadata), utc_now(), intake_category, declared_document_type),
    )
    upload_id = int(cursor.lastrowid)
    sql.commit(); sql.close()

    if not extraction:
        if contextual_source:
            _emit_progress(progress_callback, "context", 72, f"Registering {detected_type.replace('_', ' ')} as company context")
            con = get_duckdb()
            record_id = f"lf-{normalise_name(detected_type)}-{digest[:16]}"
            con.execute(
                "INSERT OR REPLACE INTO generic_documents(id, document_type, title, record_json, source_file, source_row) VALUES (?, ?, ?, ?, ?, 1)",
                (record_id, detected_type, filename, json.dumps({"text": text[:12000], "extraction_method": extraction_method}, ensure_ascii=False), filename),
            )
            con.close()
            data_version = next_data_version(COMPANY_ID, f"Store contextual PDF {filename}", [detected_type])
            pipeline = refresh_context_layers([detected_type]) if detected_type in CONTEXT_ONLY_DATASETS else refresh_gold_layers([detected_type])
            sql = get_sqlite()
            sql.execute(
                "UPDATE uploaded_files SET rows_imported=1, row_count=1, rows_new=1, data_version=?, baseline_version=?, mapping_confidence=?, processing_status='stored_source', mapping_status='mapped', metadata_json=?, last_processed_at=? WHERE id=?",
                (data_version, int(pipeline.get("baseline_version", 0)), 0.96, json.dumps(metadata, default=str), utc_now(), upload_id),
            )
            sql.commit(); sql.close()
            _emit_progress(progress_callback, "context", 92, "Updating company context without posting financial values")
            return {
                "duplicate": False,
                "upload_id": upload_id,
                "file_id": file_id,
                "filename": filename,
                "document_type": detected_type,
                "intake_category": intake_category,
                "rows_imported": 1,
                "columns": ["document_text"],
                "issues": ["The document was stored as company context. No financial values were posted from prose text."],
                "needs_mapping": False,
                "processing": {"new": 1, "changed": 0, "unchanged": 0, "rejected": 0},
                "data_version": data_version,
                "baseline_version": pipeline.get("baseline_version", 0),
                "affected_metrics": pipeline.get("affected_metrics", []),
                "storage": f"Bronze PDF + {extraction_method} text + company context",
            }

        _emit_progress(progress_callback, "review", 72, "No reliable structured fields found; preparing a validation task")
        con = get_duckdb()
        task_id = f"task-doc-{digest[:16]}"
        con.execute(
            "INSERT OR REPLACE INTO account_validation_tasks(id, task_type, source_id, source_file, counterparty, description, amount, suggested_account_code, suggested_account_name, suggested_tax_code, confidence, reason, status, created_at, resolved_at) VALUES (?, 'document_extraction', ?, ?, '', ?, 0, '', '', 'REVIEW', ?, ?, 'open', ?, NULL)",
            (task_id, str(upload_id), filename, "PDF requires field validation", 0.2 if not text else 0.45, "No reliable invoice structure could be extracted. A scanned PDF may require OCR.", utc_now()),
        )
        con.close()
        _emit_progress(progress_callback, "context", 92, "Registering the source for agent context and human review")
        return {
            "duplicate": False,
            "upload_id": upload_id,
            "file_id": file_id,
            "filename": filename,
            "document_type": detected_type,
            "intake_category": intake_category,
            "rows_imported": 0,
            "columns": [],
            "issues": [f"The PDF was preserved and made readable to the agent where text was available ({extraction_method}). Structured fields still require mapping or human validation before they affect financial metrics."],
            "needs_mapping": False,
            "processing": {"new": 0, "changed": 0, "unchanged": 0, "rejected": 1},
            "data_version": 0,
            "baseline_version": current_baseline_version(COMPANY_ID),
            "affected_metrics": [],
            "storage": "Bronze source evidence + validation queue",
        }

    document_type, record, confidence = extraction
    _emit_progress(progress_callback, "mapping", 55, f"Mapped the PDF to {document_type.replace('_', ' ')} fields")
    frame = pl.DataFrame([record], strict=False)
    mapping = suggested_mapping(frame.columns, document_type)
    silver_path = _write_silver(frame, document_type, file_id, "PDF extracted invoice")
    _emit_progress(progress_callback, "committing", 68, "Writing verified extracted fields to the business database")
    counts, issues = import_known_rows(frame, document_type, filename, mapping, upload_id, "PDF extracted invoice")
    affected = [document_type]
    data_version = next_data_version(COMPANY_ID, f"Extract PDF invoice {filename}", affected) if counts["new"] + counts["changed"] else 0
    _emit_progress(progress_callback, "analytics", 80, "Refreshing accounting, dashboards and Gold analytics")
    pipeline = refresh_gold_layers(affected) if data_version else {}
    from .accounting import rebuild_accounting_from_sources
    accounting = rebuild_accounting_from_sources()
    metadata.update({"extracted_record": record, "mapping": mapping, "confidence": confidence, "accounting": accounting})
    sql = get_sqlite()
    sql.execute(
        "UPDATE uploaded_files SET document_type=?, rows_imported=?, curated_path=?, columns_json=?, mapping_status='mapped', silver_paths_json=?, processing_status='committed', mapping_confidence=?, row_count=1, rows_new=?, rows_changed=?, rows_unchanged=?, rows_rejected=?, data_version=?, baseline_version=?, metadata_json=?, last_processed_at=? WHERE id=?",
        (document_type, counts["new"] + counts["changed"], str(silver_path), json.dumps(frame.columns), json.dumps([str(silver_path)]), confidence, counts["new"], counts["changed"], counts["unchanged"], counts["rejected"], data_version, int(pipeline.get("baseline_version", 0)), json.dumps(metadata, default=str), utc_now(), upload_id),
    )
    sql.commit(); sql.close()
    _emit_progress(progress_callback, "context", 93, "Preparing the file-specific agent explanation")
    return {
        "duplicate": False,
        "upload_id": upload_id,
        "file_id": file_id,
        "filename": filename,
        "document_type": document_type,
        "intake_category": intake_category,
        "detected_document_types": [document_type],
        "rows_imported": counts["new"] + counts["changed"],
        "columns": frame.columns,
        "issues": issues + (["Invoice fields were extracted automatically, but categorisation still needs human review."] if accounting.get("items_needing_review") else []),
        "needs_mapping": False,
        "mapping_requests": [],
        "processing": counts,
        "data_version": data_version,
        "baseline_version": pipeline.get("baseline_version", 0),
        "affected_metrics": pipeline.get("affected_metrics", []),
        "storage": f"Bronze PDF + {extraction_method} extraction + Silver invoice + accounting ledger",
    }


def process_upload(
    filename: str, content: bytes, intake_category: str = "recurring", declared_document_type: str = "auto",
    progress_callback: ProgressCallback | None = None,
) -> dict[str, Any]:
    _emit_progress(progress_callback, "identifying", 10, "Fingerprinting and identifying the workbook")
    digest = hashlib.sha256(content).hexdigest(); file_id = f"file_{digest[:16]}"; job_id = f"job_{uuid.uuid4().hex[:16]}"
    sql = get_sqlite()
    duplicate = sql.execute("SELECT * FROM uploaded_files WHERE sha256=?", (digest,)).fetchone()
    if duplicate:
        sql.close()
        duplicate, corrective_retry, category_changed = _prepare_existing_duplicate(duplicate, intake_category, declared_document_type)
        sql = get_sqlite()
        if duplicate:
            result = {
                "duplicate": True,
                "category_changed": category_changed,
                "corrective_retry": corrective_retry,
                "upload_id": duplicate["id"],
                "filename": filename,
                "document_type": duplicate["document_type"],
                "intake_category": duplicate["intake_category"] or intake_category,
                "rows_imported": int(duplicate["rows_imported"] or 0),
                "issues": ["This exact file was already uploaded. No duplicate business rows were created."],
                "needs_mapping": str(duplicate["mapping_status"] or "") == "needs_mapping",
                "columns": json.loads(duplicate["columns_json"] or "[]"),
                "processing": {"new": 0, "changed": 0, "unchanged": int(duplicate["row_count"] or 0)},
                "data_version": int(duplicate["data_version"] or 0),
                "baseline_version": int(duplicate["baseline_version"] or 0),
            }
            sql.close()
            return result
        sql.close()
        sql = get_sqlite()
    _emit_progress(progress_callback, "preserving", 20, "Preserving the original file in the Bronze evidence layer")
    original, metadata_path = _write_bronze(file_id, filename, content, digest)
    cursor = sql.execute(
        "INSERT INTO uploaded_files(created_at, filename, sha256, document_type, rows_imported, curated_path, columns_json, mapping_status, company_id, file_id, original_filename, file_type, file_size, raw_path, bronze_path, silver_paths_json, processing_status, metadata_json, last_processed_at, intake_category, declared_document_type) VALUES (?, ?, ?, 'profiling', 0, '', '[]', 'profiling', ?, ?, ?, ?, ?, ?, ?, '[]', 'profiling', '{}', ?, ?, ?)",
        (utc_now(), filename, digest, COMPANY_ID, file_id, filename, Path(filename).suffix.lower(), len(content), str(original), str(original.parent), utc_now(), intake_category, declared_document_type),
    )
    upload_id = int(cursor.lastrowid)
    sql.execute("INSERT INTO import_jobs(job_id, company_id, upload_id, started_at, status, stage) VALUES (?, ?, ?, ?, 'running', 'profiling')", (job_id, COMPANY_ID, upload_id, utc_now())); sql.commit(); sql.close()
    try:
        _emit_progress(progress_callback, "extracting", 32, "Reading sheets, columns and business records")
        frames = read_workbook(original)
        if not frames or sum(frame.height for frame in frames.values()) == 0: raise ValueError("The uploaded file contains no data rows.")
        all_types: list[str] = []; all_columns: set[str] = set(); silver_paths: list[str] = []; issues: list[str] = []
        totals = {"new": 0, "changed": 0, "unchanged": 0, "rejected": 0}; mapping_needed: list[dict[str, Any]] = []
        non_empty_frames = [item for item in frames.items() if item[1].height > 0]
        for sheet_index, (sheet_name, frame) in enumerate(non_empty_frames, start=1):
            if frame.height == 0: continue
            _emit_progress(progress_callback, "mapping", 38 + int(16 * sheet_index / max(len(non_empty_frames), 1)), f"Identifying fields in {sheet_name}")
            all_columns.update(frame.columns)
            detected_type, confidence, auto_mapping = detect_document_type(frame.columns, filename, sheet_name)
            doc_type = declared_document_type if declared_document_type != "auto" else detected_type
            if declared_document_type != "auto" and declared_document_type != detected_type:
                _, declared_confidence, declared_mapping = detect_document_type(frame.columns, f"{declared_document_type}.csv", declared_document_type)
                auto_mapping = declared_mapping or auto_mapping
                confidence = max(confidence, declared_confidence, 0.72)
            signature = schema_signature(frame.columns)
            saved_mapping = load_mapping_profile(doc_type, signature)
            mapping = saved_mapping or auto_mapping
            silver_path = _write_silver(frame, doc_type, file_id, sheet_name); silver_paths.append(str(silver_path)); all_types.append(doc_type)
            required = REQUIRED_FIELDS.get(doc_type, [])
            can_commit = doc_type != "generic" and all(mapping.get(field) for field in required)
            if can_commit:
                counts, row_issues = import_known_rows(frame, doc_type, filename, mapping, upload_id, sheet_name)
                for key in totals: totals[key] += counts[key]
                issues.extend(row_issues)
                if saved_mapping is None and confidence >= 0.82: save_mapping_profile(doc_type, signature, mapping)
            else:
                totals["rejected"] += frame.height
                mapping_needed.append({"sheet": sheet_name, "document_type": doc_type, "confidence": confidence, "columns": frame.columns, "rows": frame.height, "suggested_mapping": mapping, "silver_path": str(silver_path), "schema_signature": signature})
        _emit_progress(progress_callback, "committing", 60, "Versioning new and changed business records")
        effective_types = sorted(set(all_types)); primary_type = effective_types[0] if len(effective_types) == 1 else "mixed_business_workbook"
        needs_mapping = bool(mapping_needed)
        affected = [item for item in effective_types if item != "generic"]
        data_version = next_data_version(COMPANY_ID, f"Import {filename}", affected, job_id) if totals["new"] + totals["changed"] > 0 else 0
        context_only = bool(affected) and set(affected).issubset(CONTEXT_ONLY_DATASETS)
        _emit_progress(
            progress_callback,
            "analytics",
            76,
            "Refreshing company and market context" if context_only else "Refreshing dependent ratios, accounting and Gold analytics",
        )
        if affected and data_version:
            pipeline = refresh_context_layers(affected) if context_only else refresh_gold_layers(affected)
        else:
            pipeline = {}
        accounting = {}
        if any(item in {"supplier_invoices", "sales_invoices", "invoices", "balance_sheet"} for item in affected):
            _emit_progress(progress_callback, "accounting", 86, "Rebuilding deterministic journals and account classifications")
            from .accounting import rebuild_accounting_from_sources
            accounting = rebuild_accounting_from_sources()
        rows_imported = totals["new"] + totals["changed"]
        metadata = {"file_metadata_path": str(metadata_path), "sheets": list(frames), "sheet_routes": mapping_needed, "detected_document_types": effective_types, "intake_category": intake_category, "declared_document_type": declared_document_type, "incremental": totals, "dependencies": {doc: DEPENDENCIES.get(doc, []) for doc in affected}, "pipeline": pipeline, "accounting": accounting}
        sql = get_sqlite()
        sql.execute(
            "UPDATE uploaded_files SET document_type=?, rows_imported=?, curated_path=?, columns_json=?, mapping_status=?, silver_paths_json=?, processing_status=?, mapping_confidence=?, row_count=?, rows_new=?, rows_changed=?, rows_unchanged=?, rows_rejected=?, data_version=?, baseline_version=?, metadata_json=?, last_processed_at=? WHERE id=?",
            (primary_type, rows_imported, silver_paths[0] if silver_paths else "", json.dumps(sorted(all_columns)), "needs_mapping" if mapping_needed else "mapped", json.dumps(silver_paths), "pending_mapping" if mapping_needed else "committed", max([item.get("confidence", 0) for item in mapping_needed] + [0.9 if affected else 0.2]), sum(frame.height for frame in frames.values()), totals["new"], totals["changed"], totals["unchanged"], totals["rejected"], data_version, int(pipeline.get("baseline_version", 0)), json.dumps(metadata, default=str), utc_now(), upload_id),
        )
        sql.execute("UPDATE import_jobs SET completed_at=?, status=?, stage=?, rows_received=?, rows_new=?, rows_changed=?, rows_unchanged=?, rows_rejected=?, affected_datasets_json=? WHERE job_id=?", (utc_now(), "pending_mapping" if mapping_needed else "completed", "mapping" if mapping_needed else "gold_refreshed", sum(frame.height for frame in frames.values()), totals["new"], totals["changed"], totals["unchanged"], totals["rejected"], json.dumps(affected), job_id)); sql.commit(); sql.close()
        if mapping_needed: issues.append("One or more sheets need confirmation before they affect business metrics. Their Silver copies are already stored safely.")
        storage = (
            "Bronze original + document-specific Silver Parquet + incremental DuckDB records + company/market context"
            if context_only
            else "Bronze original + document-specific Silver Parquet + incremental DuckDB records + Gold decision features"
        )
        return {"duplicate": False, "upload_id": upload_id, "file_id": file_id, "filename": filename, "document_type": primary_type, "intake_category": intake_category, "detected_document_types": effective_types, "rows_imported": rows_imported, "columns": sorted(all_columns), "issues": issues[:50], "needs_mapping": needs_mapping, "mapping_requests": mapping_needed, "processing": totals, "data_version": data_version, "baseline_version": pipeline.get("baseline_version", 0), "affected_metrics": pipeline.get("affected_metrics", []), "storage": storage}
    except Exception as exc:
        sql = get_sqlite(); sql.execute("UPDATE uploaded_files SET processing_status='failed', mapping_status='failed', last_processed_at=? WHERE id=?", (utc_now(), upload_id)); sql.execute("UPDATE import_jobs SET completed_at=?, status='failed', stage='failed', error_message=? WHERE job_id=?", (utc_now(), str(exc), job_id)); sql.commit(); sql.close(); raise


def apply_manual_mapping(upload_id: int, document_type: str, mapping: dict[str, str]) -> dict[str, Any]:
    sql = get_sqlite(); row = sql.execute("SELECT * FROM uploaded_files WHERE id=?", (upload_id,)).fetchone()
    if not row: sql.close(); raise ValueError("Upload not found.")
    metadata = json.loads(row["metadata_json"] or "{}")
    requests = metadata.get("sheet_routes") or []
    if not requests:
        # Backward-compatible single-file mapping.
        requests = [{"sheet": "CSV", "silver_path": row["curated_path"], "schema_signature": schema_signature(json.loads(row["columns_json"] or "[]"))}]
    target = requests[0]; path = Path(target.get("silver_path") or row["curated_path"])
    if not path.exists(): sql.close(); raise ValueError("The stored Silver Parquet file is missing.")
    frame = pl.read_parquet(path); counts, issues = import_known_rows(frame, document_type, row["filename"], mapping, upload_id, str(target.get("sheet") or "CSV"))
    signature = str(target.get("schema_signature") or schema_signature(frame.columns)); save_mapping_profile(document_type, signature, mapping)
    affected = [document_type]; data_version = next_data_version(COMPANY_ID, f"Manual mapping for {row['filename']}", affected) if counts["new"] + counts["changed"] > 0 else int(row["data_version"] or 0)
    if counts["new"] + counts["changed"] > 0:
        pipeline = refresh_context_layers(affected) if set(affected).issubset(CONTEXT_ONLY_DATASETS) else refresh_gold_layers(affected)
    else:
        pipeline = {}
    if document_type in {"supplier_invoices", "sales_invoices", "invoices"}:
        from .accounting import rebuild_accounting_from_sources
        rebuild_accounting_from_sources()
    remaining = requests[1:]
    metadata["sheet_routes"] = remaining; metadata.setdefault("manual_mappings", []).append({"document_type": document_type, "mapping": mapping, "applied_at": utc_now(), "counts": counts})
    rows_imported = int(row["rows_imported"] or 0) + counts["new"] + counts["changed"]
    sql.execute("UPDATE uploaded_files SET document_type=?, rows_imported=?, mapping_status=?, processing_status=?, rows_new=rows_new+?, rows_changed=rows_changed+?, rows_unchanged=rows_unchanged+?, rows_rejected=?, data_version=?, baseline_version=?, metadata_json=?, last_processed_at=? WHERE id=?", (document_type if not remaining else "mixed_business_workbook", rows_imported, "needs_mapping" if remaining else "mapped", "pending_mapping" if remaining else "committed", counts["new"], counts["changed"], counts["unchanged"], sum(int(item.get("rows", 0)) for item in remaining), data_version, int(pipeline.get("baseline_version", row["baseline_version"] or 0)), json.dumps(metadata, default=str), utc_now(), upload_id)); sql.commit(); sql.close()
    return {
        "upload_id": upload_id,
        "document_type": document_type,
        "rows_imported": rows_imported,
        "issues": issues,
        "needs_mapping": bool(remaining),
        "mapping_requests": remaining,
        "columns": remaining[0].get("columns", []) if remaining else frame.columns,
        "processing": counts,
        "data_version": data_version,
        "baseline_version": pipeline.get("baseline_version", 0),
        "affected_metrics": pipeline.get("affected_metrics", []),
    }
